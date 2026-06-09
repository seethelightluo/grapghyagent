"""Common read-only tools shared by graph agents and node execution."""
from __future__ import annotations

import csv
import fnmatch
import json
import os
import shutil
from pathlib import Path
from typing import Any

from .context_budget import clip_text


TEXT_SUFFIXES = {
    ".c",
    ".cfg",
    ".conf",
    ".cpp",
    ".css",
    ".csv",
    ".h",
    ".html",
    ".ini",
    ".ipynb",
    ".java",
    ".js",
    ".json",
    ".jsonl",
    ".log",
    ".md",
    ".py",
    ".r",
    ".rs",
    ".sh",
    ".sql",
    ".toml",
    ".ts",
    ".tsx",
    ".txt",
    ".xml",
    ".yaml",
    ".yml",
}

TABLE_SUFFIXES = {".csv", ".tsv", ".xlsx", ".xlsm", ".xls"}
PDF_SUFFIXES = {".pdf"}
IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tif", ".tiff"}


def read_file(
    file_path: str | os.PathLike[str],
    *,
    max_chars: int = 120_000,
    offset: int = 0,
    encoding: str = "utf-8",
    workspace_root: str | os.PathLike[str] | None = None,
    allow_outside_workspace: bool = True,
) -> str:
    path = resolve_path(
        file_path,
        workspace_root=workspace_root,
        allow_outside_workspace=allow_outside_workspace,
    )
    text = path.read_text(encoding=encoding, errors="replace")
    if offset > 0:
        text = text[offset:]
    return clip_text(text, max_chars, label=path.name)


def read_file_content(
    file_path: str | os.PathLike[str],
    *,
    max_chars: int = 120_000,
    page_range: str | None = None,
    sheet: str | None = None,
    max_rows: int = 80,
    workspace_root: str | os.PathLike[str] | None = None,
    allow_outside_workspace: bool = True,
) -> str:
    path = resolve_path(
        file_path,
        workspace_root=workspace_root,
        allow_outside_workspace=allow_outside_workspace,
    )
    suffix = path.suffix.lower()
    if suffix in PDF_SUFFIXES:
        return read_pdf(path, page_range=page_range, max_chars=max_chars)
    if suffix in TABLE_SUFFIXES:
        return read_table_file(path, sheet=sheet, max_rows=max_rows, max_chars=max_chars)
    if suffix in IMAGE_SUFFIXES:
        return read_image_ocr(path, max_chars=max_chars)
    if suffix in TEXT_SUFFIXES or _looks_like_text(path):
        return read_file(path, max_chars=max_chars, workspace_root=workspace_root, allow_outside_workspace=allow_outside_workspace)
    return _binary_file_summary(path)


def summarize_file_for_prompt(
    file_record: dict[str, Any] | str | os.PathLike[str],
    *,
    max_chars: int = 120_000,
    workspace_root: str | os.PathLike[str] | None = None,
) -> str:
    if isinstance(file_record, dict):
        name = str(file_record.get("name") or file_record.get("path") or file_record.get("storage_path") or "file")
        path = file_record.get("storage_path") or file_record.get("path") or file_record.get("uri")
    else:
        path = file_record
        name = Path(str(file_record)).name
    if not path:
        return f"### 文件：{name}\n无法读取：缺少文件路径。"
    try:
        content = read_file_content(path, max_chars=max_chars, workspace_root=workspace_root)
    except Exception as exc:  # noqa: BLE001
        content = f"无法读取：{exc}"
    return f"### 文件：{name}\n```text\n{content}\n```"


def read_pdf(
    file_path: str | os.PathLike[str],
    *,
    page_range: str | None = None,
    max_chars: int = 120_000,
    workspace_root: str | os.PathLike[str] | None = None,
    allow_outside_workspace: bool = True,
) -> str:
    path = resolve_path(
        file_path,
        workspace_root=workspace_root,
        allow_outside_workspace=allow_outside_workspace,
    )
    pages = _parse_page_range(page_range)
    try:
        from pypdf import PdfReader  # type: ignore
    except Exception:
        try:
            from PyPDF2 import PdfReader  # type: ignore
        except Exception:
            return (
                f"PDF 文件：{path.name}\n"
                "当前环境未安装 pypdf/PyPDF2，无法抽取正文。"
            )
    reader = PdfReader(str(path))
    page_count = len(reader.pages)
    selected = pages or range(page_count)
    parts = [f"PDF 文件：{path.name}，页数：{page_count}"]
    for index in selected:
        if index < 0 or index >= page_count:
            continue
        try:
            text = reader.pages[index].extract_text() or ""
        except Exception as exc:  # noqa: BLE001
            text = f"[第 {index + 1} 页抽取失败：{exc}]"
        parts.append(f"\n--- 第 {index + 1} 页 ---\n{text.strip()}")
    return clip_text("\n".join(parts), max_chars, label=path.name)


def read_image_ocr(
    file_path: str | os.PathLike[str],
    *,
    max_chars: int = 60_000,
    workspace_root: str | os.PathLike[str] | None = None,
    allow_outside_workspace: bool = True,
) -> str:
    path = resolve_path(
        file_path,
        workspace_root=workspace_root,
        allow_outside_workspace=allow_outside_workspace,
    )
    metadata = _image_metadata(path)
    try:
        import pytesseract  # type: ignore
        from PIL import Image  # type: ignore
    except Exception:
        return (
            f"图片文件：{path.name}\n"
            f"{metadata}\n"
            "当前环境未安装 PIL/pytesseract，无法 OCR。"
        ).strip()
    try:
        with Image.open(path) as image:
            text = pytesseract.image_to_string(image)
    except Exception as exc:  # noqa: BLE001
        return f"图片文件：{path.name}\n{metadata}\nOCR 失败：{exc}".strip()
    return clip_text(f"图片文件：{path.name}\n{metadata}\n\nOCR 文本：\n{text}", max_chars, label=path.name)


def read_table_file(
    file_path: str | os.PathLike[str],
    *,
    sheet: str | None = None,
    max_rows: int = 80,
    max_chars: int = 120_000,
    workspace_root: str | os.PathLike[str] | None = None,
    allow_outside_workspace: bool = True,
) -> str:
    path = resolve_path(
        file_path,
        workspace_root=workspace_root,
        allow_outside_workspace=allow_outside_workspace,
    )
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        rows = _read_delimited_rows(path, delimiter=delimiter, max_rows=max_rows)
        return clip_text(_format_table(path.name, rows, total_note=None), max_chars, label=path.name)
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception:
            return (
                f"表格文件：{path.name}\n"
                "当前环境未安装 openpyxl，无法读取 Excel。"
            )
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook[sheet] if sheet and sheet in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        rows: list[list[str]] = []
        for row_index, row in enumerate(worksheet.iter_rows(values_only=True)):
            if row_index >= max_rows:
                break
            rows.append(["" if value is None else str(value) for value in row])
        total_note = f"sheet={worksheet.title}, shown_rows<={max_rows}"
        return clip_text(_format_table(path.name, rows, total_note=total_note), max_chars, label=path.name)
    return read_file(path, max_chars=max_chars)


def glob_files(
    pattern: str,
    *,
    path: str | os.PathLike[str] | None = None,
    max_results: int = 200,
    workspace_root: str | os.PathLike[str] | None = None,
    allow_outside_workspace: bool = True,
) -> list[dict[str, Any]]:
    root = resolve_path(
        path or workspace_root or ".",
        workspace_root=workspace_root,
        allow_outside_workspace=allow_outside_workspace,
    )
    if root.is_file():
        root = root.parent
    matches: list[dict[str, Any]] = []
    for item in root.rglob(pattern):
        if len(matches) >= max_results:
            break
        matches.append(_file_info(item))
    return matches


def grep_files(
    pattern: str,
    *,
    path: str | os.PathLike[str] | None = None,
    glob: str | None = None,
    case_insensitive: bool = False,
    max_results: int = 200,
    workspace_root: str | os.PathLike[str] | None = None,
    allow_outside_workspace: bool = True,
) -> list[dict[str, Any]]:
    root = resolve_path(
        path or workspace_root or ".",
        workspace_root=workspace_root,
        allow_outside_workspace=allow_outside_workspace,
    )
    if _has_rg():
        rg_result = _grep_with_rg(
            pattern,
            root,
            glob=glob,
            case_insensitive=case_insensitive,
            max_results=max_results,
        )
        if rg_result is not None:
            return rg_result
    return _grep_python(
        pattern,
        root,
        glob=glob,
        case_insensitive=case_insensitive,
        max_results=max_results,
    )


def resolve_path(
    file_path: str | os.PathLike[str],
    *,
    workspace_root: str | os.PathLike[str] | None = None,
    allow_outside_workspace: bool = True,
) -> Path:
    root = Path(workspace_root).expanduser().resolve() if workspace_root else None
    path = Path(file_path).expanduser()
    if not path.is_absolute() and root:
        path = root / path
    resolved = path.resolve()
    if root and not allow_outside_workspace and not _is_relative_to(resolved, root):
        raise PermissionError(f"path is outside workspace: {resolved}")
    if not resolved.exists():
        raise FileNotFoundError(str(resolved))
    return resolved


def _read_delimited_rows(path: Path, *, delimiter: str, max_rows: int) -> list[list[str]]:
    rows: list[list[str]] = []
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        for index, row in enumerate(reader):
            if index >= max_rows:
                break
            rows.append([str(value) for value in row])
    return rows


def _format_table(name: str, rows: list[list[str]], *, total_note: str | None) -> str:
    if not rows:
        return f"表格文件：{name}\n无可展示行。"
    widths = [0] * max(len(row) for row in rows)
    for row in rows:
        for index, value in enumerate(row):
            widths[index] = min(48, max(widths[index], len(value)))
    lines = [f"表格文件：{name}"]
    if total_note:
        lines.append(total_note)
    for row in rows:
        padded = [
            (value[:45] + "...") if len(value) > 48 else value
            for value in row
        ]
        lines.append(" | ".join(value.ljust(widths[index]) for index, value in enumerate(padded)))
    return "\n".join(lines)


def _parse_page_range(value: str | None) -> range | list[int] | None:
    text = str(value or "").strip()
    if not text:
        return None
    pages: list[int] = []
    for part in text.split(","):
        item = part.strip()
        if not item:
            continue
        if "-" in item:
            start_text, end_text = item.split("-", 1)
            try:
                start = max(1, int(start_text))
                end = max(start, int(end_text))
            except ValueError:
                continue
            pages.extend(range(start - 1, end))
        else:
            try:
                pages.append(max(1, int(item)) - 1)
            except ValueError:
                continue
    return sorted(set(pages)) if pages else None


def _image_metadata(path: Path) -> str:
    try:
        from PIL import Image  # type: ignore
    except Exception:
        return f"大小：{path.stat().st_size} bytes"
    try:
        with Image.open(path) as image:
            return f"尺寸：{image.width}x{image.height}，模式：{image.mode}，大小：{path.stat().st_size} bytes"
    except Exception:
        return f"大小：{path.stat().st_size} bytes"


def _binary_file_summary(path: Path) -> str:
    return (
        f"二进制或暂不支持的文件：{path.name}\n"
        f"扩展名：{path.suffix or '(none)'}\n"
        f"大小：{path.stat().st_size} bytes"
    )


def _looks_like_text(path: Path) -> bool:
    try:
        sample = path.read_bytes()[:4096]
    except OSError:
        return False
    if b"\x00" in sample:
        return False
    try:
        sample.decode("utf-8")
        return True
    except UnicodeDecodeError:
        return False


def _file_info(path: Path) -> dict[str, Any]:
    stat = path.stat()
    return {
        "path": str(path),
        "name": path.name,
        "is_file": path.is_file(),
        "size": stat.st_size if path.is_file() else None,
    }


def _has_rg() -> bool:
    return bool(shutil.which("rg"))


def _grep_with_rg(
    pattern: str,
    root: Path,
    *,
    glob: str | None,
    case_insensitive: bool,
    max_results: int,
) -> list[dict[str, Any]] | None:
    import subprocess

    command = ["rg", "--json", "--line-number", "--max-count", str(max_results)]
    if case_insensitive:
        command.append("--ignore-case")
    if glob:
        command.extend(["--glob", glob])
    command.append(pattern)
    command.append(str(root))
    try:
        completed = subprocess.run(command, capture_output=True, text=True, encoding="utf-8", errors="replace", timeout=30)
    except Exception:
        return None
    if completed.returncode not in {0, 1}:
        return None
    results: list[dict[str, Any]] = []
    for line in completed.stdout.splitlines():
        if len(results) >= max_results:
            break
        try:
            item = json.loads(line)
        except json.JSONDecodeError:
            continue
        if item.get("type") != "match":
            continue
        data = item.get("data") or {}
        results.append({
            "path": (data.get("path") or {}).get("text"),
            "line_number": data.get("line_number"),
            "line": (data.get("lines") or {}).get("text", "").rstrip("\n"),
        })
    return results


def _grep_python(
    pattern: str,
    root: Path,
    *,
    glob: str | None,
    case_insensitive: bool,
    max_results: int,
) -> list[dict[str, Any]]:
    needle = pattern.lower() if case_insensitive else pattern
    files = [root] if root.is_file() else [item for item in root.rglob("*") if item.is_file()]
    results: list[dict[str, Any]] = []
    for file_path in files:
        if len(results) >= max_results:
            break
        if glob and not fnmatch.fnmatch(file_path.name, glob):
            continue
        if file_path.suffix.lower() not in TEXT_SUFFIXES and not _looks_like_text(file_path):
            continue
        try:
            with file_path.open("r", encoding="utf-8", errors="replace") as handle:
                for line_number, line in enumerate(handle, start=1):
                    haystack = line.lower() if case_insensitive else line
                    if needle in haystack:
                        results.append({
                            "path": str(file_path),
                            "line_number": line_number,
                            "line": line.rstrip("\n"),
                        })
                        if len(results) >= max_results:
                            break
        except OSError:
            continue
    return results


def _is_relative_to(path: Path, root: Path) -> bool:
    try:
        path.relative_to(root)
        return True
    except ValueError:
        return False


__all__ = [
    "glob_files",
    "grep_files",
    "read_file",
    "read_file_content",
    "read_image_ocr",
    "read_pdf",
    "read_table_file",
    "resolve_path",
    "summarize_file_for_prompt",
]
